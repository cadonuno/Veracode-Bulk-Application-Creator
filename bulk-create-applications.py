import sys
import requests
import getopt
import json
import urllib.parse
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
import openpyxl
import time
import xml.etree.ElementTree as ET  # for parsing XML

from veracode_api_signing.credentials import get_credentials

class NoExactMatchFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message



json_headers = {
    "User-Agent": "Bulk application creation - python script",
    "Content-Type": "application/json"
}

xml_headers = {
    "User-Agent": "Bulk application creation - python script",
    "Content-Type": "application/xml"
}

failed_attempts = 0
max_attempts_per_request = 10
sleep_time = 10

last_column = 0
non_custom_field_headers={"Application Name",
                            "Business Criticality",
                            "Policy",
                            "Submitting party", #not used, here just so it gets ignored
                            "Vendor", #not used, here just so it gets ignored
                            "Origin",
                            "Industry",
                            "Application Purpose",
                            "Deployment Method",
                            "Description",
                            "Tags",
                            "Business Unit",
                            "Business Owner",
                            "Owner Email",
                            "Teams",
                            "Dynamic Scan Approval",
                            "Archer Application Name"}

def print_help():
    """Prints command line options and exits"""
    print("""bulk-create-applications.py -f <excel_file_with_application_definitions> -r <header_row> [-d]"
        Reads all lines in <excel_file_with_application_definitions>, for each line, it will create a new application profile
        <header_row> defines which row contains your table headers, which will be read to determine where each field goes.
""")
    sys.exit()

def request_encode(value_to_encode):
    return urllib.parse.quote(value_to_encode, safe='')

def find_exact_match(list, to_find, field_name):
    for index in range(len(list)):
        if list[index][field_name] == to_find:
            return list[index]
    print(f"Unable to find a member of list with {field_name} equal to {to_find}")
    raise NoExactMatchFoundException(f"Unable to find a member of list with {field_name} equal to {to_find}")

def get_field_value(excel_headers, excel_sheet, row, field_header):
    field_to_get = field_header.strip()
    if field_to_get in excel_headers:
        field_value = excel_sheet.cell(row = row, column = excel_headers[field_to_get]).value
        if field_value:
            return field_value
    return ""

def get_business_owners(excel_headers, excel_sheet, row):
    name=get_field_value(excel_headers, excel_sheet, row, "Business Owner")
    email=get_field_value(excel_headers, excel_sheet, row, "Owner Email")
    if not name or not email:
        return ""
    return f''',
    "business_owners": [
      {{
        "email": "{email}",
        "name": "{name}"
      }}
    ]'''

def get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, field_to_check, field_to_get, is_exact_match, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}{api_to_call}"
    if verbose:
        print(f"Calling: {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers)
    data = response.json()

    if response.status_code == 200:
        if verbose:
            print(data)
        if "_embedded" in data and len(data["_embedded"][list_name]) > 0:
            return (find_exact_match(data["_embedded"][list_name], item_to_find, field_to_check) if is_exact_match else data["_embedded"][list_name][0])[field_to_get]
        else:
            print(f"ERROR: No {list_name} named '{item_to_find}' found")
            return f"ERROR: No {list_name} named '{item_to_find}' found"
    else:
        print(f"ERROR: trying to get {list_name} named {item_to_find}")
        print(f"ERROR: code: {response.status_code}")
        print(f"ERROR: value: {data}")
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, field_to_check, field_to_get, verbose)
        else:
            return f"ERROR: trying to get {list_name} named {item_to_find}"

def get_business_unit(api_base, excel_headers, excel_sheet, row, verbose):
    business_unit_name = get_field_value(excel_headers, excel_sheet, row, "Business Unit")
    if not business_unit_name:
        return ""
    else:
        return f''',
    "business_unit": {{
      "guid": "{get_item_from_api_call(api_base, "api/authn/v2/business_units?bu_name="+ request_encode(business_unit_name), business_unit_name, "business_units", "bu_name", "bu_id", True, verbose)}"
    }}'''

def get_policy(api_base, excel_headers, excel_sheet, row, verbose):
    policy_name = get_field_value(excel_headers, excel_sheet, row, "Policy")
    if not policy_name:
        return ""
    else:
        return f''',
    "policies": [{{
      "guid": "{get_item_from_api_call(api_base, "appsec/v1/policies?category=APPLICATION&name_exact=true&public_policy=true&name="+ request_encode(policy_name), policy_name, "policy_versions", "name", "guid", False, verbose)}"
    }}]'''

def get_team_value(api_base, team_name, verbose):
    team_guid = get_item_from_api_call(api_base, "api/authn/v2/teams?all_for_org=true&team_name="+ request_encode(team_name), team_name, "teams", "team_name", "team_id", True, verbose)
    if team_guid:
        return f'''{{
        "guid": "{team_guid}"
      }}'''
    else:
        return ""

def get_teams(api_base, excel_headers, excel_sheet, row, verbose):
    all_teams_base = get_field_value(excel_headers, excel_sheet, row, "Teams")
    if not all_teams_base:
        return ""
    all_teams = all_teams_base.split(",")
    inner_team_list = ""
    for team_name in all_teams:
        team_value = get_team_value(api_base, team_name.strip(), verbose)
        if team_name: 
            inner_team_list = inner_team_list + (""",
            """ if inner_team_list else "") + team_value
    if inner_team_list:
        return f''',
            "teams": [
                {inner_team_list}
            ]'''
    else:
        return ""
        
def get_application_settings(excel_headers, excel_sheet, row):
    base_value = get_field_value(excel_headers, excel_sheet, row, "Dynamic Scan Approval")
    value = False
    if base_value:
        value = str(base_value).strip().lower() == "false"
    return f''',
    "settings": {{
      "dynamic_scan_approval_not_required": {str(value).lower()}
    }}'''

def get_custom_fields(excel_headers, excel_sheet, row):
    global non_custom_field_headers
    inner_custom_fields_list = ""
    for field in excel_headers:
        if not field in non_custom_field_headers:
            value = excel_sheet.cell(row = row, column=excel_headers[field]).value
            if value:
                found_field = f'''{{
                    "name": "{field}",
                    "value": "{value}"
                }}'''
                inner_custom_fields_list = inner_custom_fields_list + (""",
                """ if inner_custom_fields_list else "") + found_field
    if inner_custom_fields_list:
        return f''',
                "custom_fields": [
                    {inner_custom_fields_list}
                ]'''
    else:
        return ""
        
def get_archer_application_name(excel_headers, excel_sheet, row):
    archer_app_name = get_field_value(excel_headers, excel_sheet, row, "Archer Application Name")
    if archer_app_name:
        return f''',
        "archer_app_name": "{archer_app_name}"'''
    else:
        return ""

def url_encode_with_plus(a_string):
    return urllib.parse.quote_plus(a_string, safe='').replace("&", "%26")

def get_error_node_value(body):
    inner_node = ET.XML(body)
    if inner_node.tag == "error" and not inner_node == None:
        return inner_node.text
    else:
        return ""
    

def set_xml_api_values(application_xml_id, api_base, excel_headers, excel_sheet, row, verbose):
    path = f"{api_base}api/5.0/updateapp.do?app_id={application_xml_id}"
    origin = get_field_value(excel_headers, excel_sheet, row, "Origin")
    industry = get_field_value(excel_headers, excel_sheet, row, "Industry")
    application_purpose = get_field_value(excel_headers, excel_sheet, row, "Application Purpose")
    deployment_method = get_field_value(excel_headers, excel_sheet, row, "Deployment Method")
    if not origin and not industry and not application_purpose and not deployment_method:
        return "success"
    print("Setting xml-only values")
    
    if origin:
        path = path + f'&origin={url_encode_with_plus(origin)}'
    if industry:
        path = path + f'&industry={url_encode_with_plus(industry)}'
    if application_purpose:
        path = path + f'&app_type={url_encode_with_plus(application_purpose)}'
    if deployment_method:
        path = path + f'&deployment_method={url_encode_with_plus(deployment_method)}'

    if verbose:
        print(path)
    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=xml_headers)
    body = response.content
    if verbose:
        print(f"status code {response.status_code}")
        if body:
            print(body)
    if response.status_code != 200:
        return f"ERROR: Unable to add xml-only fields to application: {response.status_code}"
    error = get_error_node_value(body)
    if not error:
        print("Successfully updated xml-only fields.")
        return "success"
    else:
        message = f"ERROR: Unable to add xml-only fields to application: {error}"
        print(message)
        return message

def create_application(api_base, excel_headers, excel_sheet, row, verbose):
    path = f"{api_base}appsec/v1/applications"
    request_content=f'''{{
        "profile": {{
            "business_criticality": "{get_field_value(excel_headers, excel_sheet, row, "Business Criticality").replace(" ", "_").upper()}"            
            {get_archer_application_name(excel_headers, excel_sheet, row)}
            {get_business_owners(excel_headers, excel_sheet, row)}
            {get_business_unit(api_base, excel_headers, excel_sheet, row, verbose)},
            "description": "{get_field_value(excel_headers, excel_sheet, row, "Description")}",
            "name": "{get_field_value(excel_headers, excel_sheet, row, "Application Name")}"
            {get_policy(api_base, excel_headers, excel_sheet, row, verbose)},
            "tags": "{get_field_value(excel_headers, excel_sheet, row, "Tags")}"
            {get_teams(api_base, excel_headers, excel_sheet, row, verbose)}
            {get_application_settings(excel_headers, excel_sheet, row)}
            {get_custom_fields(excel_headers, excel_sheet, row)}
        }}
    }}'''
    if verbose:
        print(request_content)

    response = requests.post(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content))

    if verbose:
        print(f"status code {response.status_code}")
        body = response.json()
        if body:
            print(body)
    if response.status_code == 200:
        print("Successfully created application profile.")
        return set_xml_api_values(response.json()["id"], api_base.replace("api", "analysiscenter"), excel_headers, excel_sheet, row, verbose)
    else:
        body = response.json()
        if (body):
            return f"Unable to create application profile: {response.status_code} - {body}"
        else:
            return f"Unable to create application profile: {response.status_code}"
    

def setup_excel_headers(excel_sheet, header_row, verbose):
    excel_headers = {}
    global last_column
    for column in range(1, excel_sheet.max_column+1):
        cell = excel_sheet.cell(row = header_row, column = column)
        if not cell or cell is None or not cell.value or cell.value.strip() == "":
            break
        to_add = cell.value
        if to_add:
            to_add = str(to_add).strip()
        if verbose:
            print(f"Adding column {column} for value {to_add}")
        excel_headers[to_add] = column
        last_column += 1
    return excel_headers

def create_all_applications(api_base, file_name, header_row, verbose):
    global failed_attempts
    excel_file = openpyxl.load_workbook(file_name)
    excel_sheet = excel_file.active
    try:
        excel_headers = setup_excel_headers(excel_sheet, header_row, verbose)
        print("Finished reading excel headers")
        if verbose:
            print("Values found are:")
            print(excel_headers)

        max_column=len(excel_headers)
        for row in range(header_row+1, excel_sheet.max_row+1):      
            failed_attempts = 0
            if verbose:
                for field in excel_headers:
                    print(f"Found column with values:")
                    print(f"{field} -> {excel_sheet.cell(row = row, column=excel_headers[field]).value}")
            status=excel_sheet.cell(row = row, column = max_column+1).value
            if (status == 'success'):
                print("Skipping row as it was already done")
            else:
                try:
                    print(f"Importing row {row-header_row}/{excel_sheet.max_row-header_row}:")
                    status = create_application(api_base, excel_headers, excel_sheet, row, verbose)
                    print(f"Finished importing row {row-header_row}/{excel_sheet.max_row-header_row}")
                    print("---------------------------------------------------------------------------")
                except NoExactMatchFoundException:
                    status= NoExactMatchFoundException.get_message()
                excel_sheet.cell(row = row, column = max_column+1).value=status
    finally:
        excel_file.save(filename=file_name)

def get_api_base():
    api_key_id, api_key_secret = get_credentials()
    api_base = "https://api.veracode.{instance}/"
    if api_key_id.startswith("vera01"):
        return api_base.replace("{instance}", "eu", 1)
    else:
        return api_base.replace("{instance}", "com", 1)

def main(argv):
    """Allows for bulk adding application profiles"""
    global failed_attempts
    global last_column
    excel_file = None
    try:
        verbose = False
        file_name = ''
        header_row = -1

        opts, args = getopt.getopt(argv, "hdf:r:", ["file_name=", "header_row="])
        for opt, arg in opts:
            if opt == '-h':
                print_help()
            if opt == '-d':
                verbose = True
            if opt in ('-f', '--file_name'):
                file_name=arg
            if opt in ('-r', '--header_row'):
                header_row=int(arg)

        api_base = get_api_base()
        if file_name and header_row> 0:
            create_all_applications(api_base, file_name, header_row, verbose)
        else:
            print_help()
    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)
    finally:
        if excel_file:
            excel_file.save(filename=file_name)


if __name__ == "__main__":
    main(sys.argv[1:])
